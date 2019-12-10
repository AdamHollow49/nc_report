import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.chart import DoughnutChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
import re

daysNC = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0}
dayTrns = {0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday', 5: 'Saturday', 6: 'Sunday'}

wbwork = load_workbook(filename="L140 NC SPREADSHEET.xlsx", data_only=True)  # find NC spreadsheet
try:
    sheetnum = wbwork.sheetnames.index('FINISHED GOODS 2020')
    sheet = wbwork.sheets[sheetnum]
except ValueError as e:
    print("Error: No sheet named 'FINISHED GOODS 2020' within workbook")


def load_Title():  # load title date to title storage array

    titleLine = []
    colTrace = 1  # counter to trace over columns
    while sheet.cell(2, colTrace).value:  # while column specified has a value, run code
        cellVal = sheet.cell(2, colTrace).value  # instantiating current cell as cellVal
        titleLine.append(cellVal)  # adding to titleLine
        colTrace += 1  # incrementing counter
    return titleLine  # return storage array


titleLine = load_Title()  # run title load
dueCol = sheet.cell(2, titleLine.index('DATE DUE') + 1).column_letter  # column letter of due date
compCol = sheet.cell(2, titleLine.index('DATE COMPLETED') + 1).column_letter  # column letter of date completed
confCol = sheet.cell(2, titleLine.index('NCGR NO.') + 1).column_letter  # column letter of NC ID
raisedCol = sheet.cell(2, titleLine.index('DATE RAISED') + 1).column_letter
dueDates = []  # due dates storage array
pendRows = []  # row number of pending NCs storage array
overRows = []  # row number of overdue NCs storage array
compCount = 0
rows = 0


def load_due_dates():  # load due dates

    rowTrace = 3  # row where due dates startwwww
    while sheet['{x}'.format(x=confCol) + str(rowTrace)].value:
        if type(sheet['{x}'.format(x=raisedCol) + str(rowTrace)].value) == datetime:
            daysNC[sheet['{x}'.format(x=raisedCol) + str(rowTrace)].value.weekday()] += 1
        dueVal = sheet['{x}'.format(x=dueCol) + str(rowTrace)].value
        dueRow = sheet['{x}'.format(x=dueCol) + str(rowTrace)].row
        compVal = sheet['{x}'.format(x=compCol) + str(rowTrace)].value
        if compVal is None or compVal == "":
            if dueVal is not None:
                dueDates.append([dueRow, dueVal])
        else:
            global compCount
            compCount += 1
            print("NCGR NO: " + sheet['{x}'.format(x=confCol) + str(rowTrace)].value + " is completed.")
        rowTrace += 1
    for x in dueDates:
        if len(str(x[1])) > 10:
            x[1] = x[1].strftime('%d/%m/%y')


def check_dates():
    today = datetime.now()
    today = datetime(today.year, today.month, today.day)
    global rows

    for dates in dueDates:
        date = dates[1]
        try:
            date = datetime(2000 + int(date[6:8]), int(date[3:5]), int(date[0:2]))
        except ValueError as e:
            print("There is a problem with the date format in row: " + dates[0])
            print("System Error", e)
        except TypeError as e:
            print("There is a problem with the date format in row: " + dates[0])
            print("System Error", e)
        delta = date - today
        if delta.days >= 0:
            pendRows.append(dates[0])
        else:
            overRows.append(dates[0])

    for rows in pendRows:
        print("NCGR NO: " + sheet['{x}'.format(x=confCol) + str(rows)].value + " is pending.")
    for rows in overRows:
        print("NCGR NO: " + sheet['{x}'.format(x=confCol) + str(rows)].value + " is overdue!")


def insert_data():
    global rows
    wbNew = Workbook()
    filePath = "nc_report.xlsx"
    newSheet = wbNew.active
    colCount = 1
    for x in range(int(len(titleLine))):
        newSheet.cell(colCount, x + 1, titleLine[x])

    # title inserted, continue with data insertion
    rowCounter = 0  # row counter
    # insert pending NCs
    for each in pendRows:
        for x in range(int(len(titleLine))):
            if type(sheet.cell(each, x + 1).value) == datetime:
                sheet.cell(each, x + 1).value = sheet.cell(each, x + 1).value.strftime('%d/%m/%y')
            newSheet.cell(pendRows.index(each) + 2, x + 1, sheet.cell(each, x + 1).value)
            if sheet.cell(each, x + 1).value == "Pending":
                newSheet.cell(rowCounter + 2, x + 1).fill = PatternFill(start_color='0070C0', end_color='000000',
                                                                        fill_type='solid')
        rowCounter += 1

    # insert overdue NCs
    for each in overRows:
        for x in range(int(len(titleLine))):
            if type(sheet.cell(each, x + 1).value) == datetime:
                sheet.cell(each, x + 1).value = sheet.cell(each, x + 1).value.strftime('%d/%m/%y')
            newSheet.cell(rowCounter + 2, x + 1, sheet.cell(each, x + 1).value)
            if sheet.cell(each, x + 1).value == "Overdue":
                newSheet.cell(rowCounter + 2, x + 1).fill = PatternFill(start_color='FF0000', end_color='000000',
                                                                        fill_type='solid')
        rowCounter += 1
        rows = rowCounter

    # set column width
    for col in newSheet.columns:  # algorithm for setting column width - credit to Ssubrat Rrudra on Stack Overflow:
        # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        maxLength = 1
        colName = re.findall('\w\d', str(col[0]))
        colName = colName[0]
        colName = re.findall('\w', str(colName))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > maxLength:
                    maxLength = len(cell.value)
            except:
                pass
        adjustedWidth = (maxLength + 6)
        newSheet.column_dimensions[colName].width = adjustedWidth
    wbNew.save(filePath)


def create_pie():
    global compCount
    wbNew = load_workbook('nc_report.xlsx')
    newSheet = wbNew.active
    data = [
        ['NC Status', 'Count'],
        ['Pending', len(pendRows)],
        ['Overdue', len(overRows)],
        ['Completed', compCount]
    ]
    newSheet.append([""])
    for row in data:
        newSheet.append(row)

    global rows
    chart = DoughnutChart()
    labels = Reference(newSheet, min_col=1, min_row=rows + 4, max_row=rows + 6)
    data = Reference(newSheet, min_col=2, min_row=rows + 3, max_row=rows + 6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "NC Report"
    chart.style = 26
    newSheet.add_chart(chart, "A{x}".format(x=rows + 7))
    wbNew.save('nc_report.xlsx')


def create_bar():
    global rows
    wbNew = load_workbook('nc_report.xlsx')
    newSheet = wbNew.active
    newSheet.append([""])
    newSheet.append(['Number of NCs', 'Day'])
    for i in range(7):
        newSheet.append([daysNC[i], dayTrns[i]])  # weird interaction where inserting a list worked but an int didn't
    wbNew.save('nc_report.xlsx')
    data = Reference(newSheet, min_col=1, min_row=rows + 8, max_row=rows + 15)
    titles = Reference(newSheet, min_col=2, min_row=rows + 9, max_row=rows + 15)
    chart = BarChart()
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(titles)
    chart.title = "NCs / Weekday"
    chart.x_axis.title = "Weekday"
    chart.y_axis.title = "No. of NCs raised"
    newSheet.add_chart(chart, "G{x}".format(x=rows + 7))
    newSheet.cell(rows + 8, 1).value = ""
    wbNew.save('nc_report.xlsx')


load_due_dates()
check_dates()
insert_data()
create_pie()
create_bar()
